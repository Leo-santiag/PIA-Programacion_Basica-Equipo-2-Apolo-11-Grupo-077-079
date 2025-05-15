# proyecto_pokemon_graficas.py

import os
import re
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage

def validar_fila(fila):
    nombre = str(fila['Nombre'])
    tipo = str(fila['Tipo'])

    if not re.fullmatch(r"[A-Za-z0-9\- ]{1,31}", nombre):
        print(f"[VALIDACIÓN] Nombre inválido: {nombre}")
        return False

    if not re.fullmatch(r"[A-Za-z]+(/[A-Za-z]+)?", tipo):
        print(f"[VALIDACIÓN] Tipo inválido: {tipo}")
        return False

    columnas_numericas = ['HP', 'Ataque', 'Defensa', 'Ataque Especial',
                          'Defensa Especial', 'Velocidad', 'Altura', 'Peso', 'Experiencia Base']
    
    for col in columnas_numericas:
        valor = fila[col]
        if pd.isnull(valor) or not isinstance(valor, (int, float)):
            print(f"[VALIDACIÓN] Valor inválido en {col} para {nombre}: {valor}")
            return False

    return True

def generar_graficas_de_barras(df):
    os.makedirs("graficas", exist_ok=True)
    os.makedirs("resultados", exist_ok=True)

    wb = Workbook()
    ws_general = wb.active
    ws_general.title = "Resumen"

    encabezados = ["ID", "Nombre", "Tipo", "HP", "Ataque", "Defensa", "Ataque Especial",
                   "Defensa Especial", "Velocidad", "Altura", "Peso", "Experiencia Base"]
    ws_general.append(encabezados)

    for _, fila in df.iterrows():
        if not validar_fila(fila):
            continue
        nombre = str(fila['Nombre'])[:31]
        tipo = fila['Tipo']

        datos = {
            'HP': fila['HP'], 'Ataque': fila['Ataque'], 'Defensa': fila['Defensa'],
            'Ataque Esp': fila['Ataque Especial'], 'Defensa Esp': fila['Defensa Especial'],
            'Velocidad': fila['Velocidad'], 'Altura': fila['Altura'], 'Peso': fila['Peso'],
            'Exp Base': fila['Experiencia Base']
        }

        plt.figure(figsize=(10, 6))
        plt.bar(datos.keys(), datos.values(), color='orange')
        plt.title(f"{tipo} - {nombre}")
        plt.xticks(rotation=45)
        plt.tight_layout()
        path = f"graficas/{nombre}_barras.png"
        plt.savefig(path)
        plt.close()

        hoja = wb.create_sheet(title=nombre)
        hoja.append(encabezados)
        hoja.append([fila[col] for col in encabezados])
        img = ExcelImage(path)
        img.anchor = 'A5'
        hoja.add_image(img)
        ws_general.append([fila[col] for col in encabezados])

    wb.save("resultados/graficas_barras.xlsx")

def generar_graficas_de_lineas(df):
    os.makedirs("graficas", exist_ok=True)
    wb = Workbook()
    ws_general = wb.active
    ws_general.title = "Resumen"
    encabezados = ["Nombre"]
    ws_general.append(encabezados)

    for _, fila in df.iterrows():
        if not validar_fila(fila):
            continue
        nombre = str(fila['Nombre'])[:31]
        valores = [
            fila['HP'], fila['Ataque'], fila['Defensa'],
            fila['Ataque Especial'], fila['Defensa Especial'], fila['Velocidad']
        ]
        etiquetas = ['HP', 'Ataque', 'Defensa', 'Atq Esp', 'Def Esp', 'Velocidad']

        plt.figure()
        plt.plot(etiquetas, valores, marker='o', color='blue')
        plt.title(f"Estadísticas de {nombre}")
        plt.xlabel("Atributos")
        plt.ylabel("Valor")
        plt.grid(True)
        plt.tight_layout()
        img_path = f"graficas/{nombre}_lineas.png"
        plt.savefig(img_path)
        plt.close()

        hoja = wb.create_sheet(title=nombre)
        hoja.append(["Gráfico de líneas para: " + nombre])
        img = ExcelImage(img_path)
        img.anchor = 'A3'
        hoja.add_image(img)
        ws_general.append([nombre])

    wb.save("resultados/graficas_lineas.xlsx")

def generar_graficas_de_pastel(df):
    os.makedirs("graficas", exist_ok=True)
    wb = Workbook()
    ws_general = wb.active
    ws_general.title = "Resumen"
    encabezados = ["Nombre"]
    ws_general.append(encabezados)

    for _, fila in df.iterrows():
        if not validar_fila(fila):
            continue
        nombre = str(fila['Nombre'])[:31]
        valores = [fila['HP'], fila['Ataque'], fila['Defensa'], fila['Velocidad']]
        etiquetas = ['HP', 'Ataque', 'Defensa', 'Velocidad']

        plt.figure()
        plt.pie(valores, labels=etiquetas, autopct='%1.1f%%', startangle=90)
        plt.title(f"Distribución Básica - {nombre}")
        plt.tight_layout()
        img_path = f"graficas/{nombre}_pastel.png"
        plt.savefig(img_path)
        plt.close()

        hoja = wb.create_sheet(title=nombre)
        hoja.append(["Gráfico de pastel para: " + nombre])
        img = ExcelImage(img_path)
        img.anchor = 'A3'
        hoja.add_image(img)
        ws_general.append([nombre])

    wb.save("resultados/graficas_pastel.xlsx")

def generar_dispersión(df):
    os.makedirs("graficas", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Dispersión"

    plt.figure(figsize=(8, 6))
    plt.scatter(df['Ataque'], df['Defensa'], c='purple', alpha=0.6)
    plt.title("Ataque vs Defensa")
    plt.xlabel("Ataque")
    plt.ylabel("Defensa")
    plt.grid(True)
    plt.tight_layout()
    img_path = "graficas/ataque_vs_defensa_dispersion.png"
    plt.savefig(img_path)
    plt.close()

    img = ExcelImage(img_path)
    img.anchor = 'A3'
    ws.append(["Gráfico de dispersión: Ataque vs Defensa"])
    ws.add_image(img)

    wb.save("resultados/grafica_dispersion.xlsx")

if __name__ == "__main__":
    archivo = "pokemones_resumen.csv"
    if not os.path.exists(archivo):
        print(f"[ERROR] No se encontró: {archivo}")
    else:
        df = pd.read_csv(archivo)
        generar_graficas_de_barras(df)
        generar_graficas_de_lineas(df)
        generar_graficas_de_pastel(df)
        generar_dispersión(df)
        print("\n Todas las gráficas y libros de Excel fueron generados correctamente en la misma carpeta.")
# proyecto_pokemon_graficas.py

import os
import re
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage

def validar_fila(fila):
    nombre = str(fila['Nombre'])
    tipo = str(fila['Tipo'])

    if not re.fullmatch(r"[A-Za-z0-9\- ]{1,31}", nombre):
        print(f"[VALIDACIÓN] Nombre inválido: {nombre}")
        return False

    if not re.fullmatch(r"[A-Za-z]+(/[A-Za-z]+)?", tipo):
        print(f"[VALIDACIÓN] Tipo inválido: {tipo}")
        return False

    columnas_numericas = ['HP', 'Ataque', 'Defensa', 'Ataque Especial',
                          'Defensa Especial', 'Velocidad', 'Altura', 'Peso', 'Experiencia Base']
    
    for col in columnas_numericas:
        valor = fila[col]
        if pd.isnull(valor) or not isinstance(valor, (int, float)):
            print(f"[VALIDACIÓN] Valor inválido en {col} para {nombre}: {valor}")
            return False

    return True

def generar_graficas_de_barras(df):
    os.makedirs("graficas", exist_ok=True)
    os.makedirs("resultados", exist_ok=True)

    wb = Workbook()
    ws_general = wb.active
    ws_general.title = "Resumen"

    encabezados = ["ID", "Nombre", "Tipo", "HP", "Ataque", "Defensa", "Ataque Especial",
                   "Defensa Especial", "Velocidad", "Altura", "Peso", "Experiencia Base"]
    ws_general.append(encabezados)

    for _, fila in df.iterrows():
        if not validar_fila(fila):
            continue
        nombre = str(fila['Nombre'])[:31]
        tipo = fila['Tipo']

        datos = {
            'HP': fila['HP'], 'Ataque': fila['Ataque'], 'Defensa': fila['Defensa'],
            'Ataque Esp': fila['Ataque Especial'], 'Defensa Esp': fila['Defensa Especial'],
            'Velocidad': fila['Velocidad'], 'Altura': fila['Altura'], 'Peso': fila['Peso'],
            'Exp Base': fila['Experiencia Base']
        }

        plt.figure(figsize=(10, 6))
        plt.bar(datos.keys(), datos.values(), color='orange')
        plt.title(f"{tipo} - {nombre}")
        plt.xticks(rotation=45)
        plt.tight_layout()
        path = f"graficas/{nombre}_barras.png"
        plt.savefig(path)
        plt.close()

        hoja = wb.create_sheet(title=nombre)
        hoja.append(encabezados)
        hoja.append([fila[col] for col in encabezados])
        img = ExcelImage(path)
        img.anchor = 'A5'
        hoja.add_image(img)
        ws_general.append([fila[col] for col in encabezados])

    wb.save("resultados/graficas_barras.xlsx")

def generar_graficas_de_lineas(df):
    os.makedirs("graficas", exist_ok=True)
    wb = Workbook()
    ws_general = wb.active
    ws_general.title = "Resumen"
    encabezados = ["Nombre"]
    ws_general.append(encabezados)

    for _, fila in df.iterrows():
        if not validar_fila(fila):
            continue
        nombre = str(fila['Nombre'])[:31]
        valores = [
            fila['HP'], fila['Ataque'], fila['Defensa'],
            fila['Ataque Especial'], fila['Defensa Especial'], fila['Velocidad']
        ]
        etiquetas = ['HP', 'Ataque', 'Defensa', 'Atq Esp', 'Def Esp', 'Velocidad']

        plt.figure()
        plt.plot(etiquetas, valores, marker='o', color='blue')
        plt.title(f"Estadísticas de {nombre}")
        plt.xlabel("Atributos")
        plt.ylabel("Valor")
        plt.grid(True)
        plt.tight_layout()
        img_path = f"graficas/{nombre}_lineas.png"
        plt.savefig(img_path)
        plt.close()

        hoja = wb.create_sheet(title=nombre)
        hoja.append(["Gráfico de líneas para: " + nombre])
        img = ExcelImage(img_path)
        img.anchor = 'A3'
        hoja.add_image(img)
        ws_general.append([nombre])

    wb.save("resultados/graficas_lineas.xlsx")

def generar_graficas_de_pastel(df):
    os.makedirs("graficas", exist_ok=True)
    wb = Workbook()
    ws_general = wb.active
    ws_general.title = "Resumen"
    encabezados = ["Nombre"]
    ws_general.append(encabezados)

    for _, fila in df.iterrows():
        if not validar_fila(fila):
            continue
        nombre = str(fila['Nombre'])[:31]
        valores = [fila['HP'], fila['Ataque'], fila['Defensa'], fila['Velocidad']]
        etiquetas = ['HP', 'Ataque', 'Defensa', 'Velocidad']

        plt.figure()
        plt.pie(valores, labels=etiquetas, autopct='%1.1f%%', startangle=90)
        plt.title(f"Distribución Básica - {nombre}")
        plt.tight_layout()
        img_path = f"graficas/{nombre}_pastel.png"
        plt.savefig(img_path)
        plt.close()

        hoja = wb.create_sheet(title=nombre)
        hoja.append(["Gráfico de pastel para: " + nombre])
        img = ExcelImage(img_path)
        img.anchor = 'A3'
        hoja.add_image(img)
        ws_general.append([nombre])

    wb.save("resultados/graficas_pastel.xlsx")

def generar_dispersión(df):
    os.makedirs("graficas", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Dispersión"

    plt.figure(figsize=(8, 6))
    plt.scatter(df['Ataque'], df['Defensa'], c='purple', alpha=0.6)
    plt.title("Ataque vs Defensa")
    plt.xlabel("Ataque")
    plt.ylabel("Defensa")
    plt.grid(True)
    plt.tight_layout()
    img_path = "graficas/ataque_vs_defensa_dispersion.png"
    plt.savefig(img_path)
    plt.close()

    img = ExcelImage(img_path)
    img.anchor = 'A3'
    ws.append(["Gráfico de dispersión: Ataque vs Defensa"])
    ws.add_image(img)

    wb.save("resultados/grafica_dispersion.xlsx")

if __name__ == "__main__":
    archivo = "pokemones_resumen.csv"
    if not os.path.exists(archivo):
        print(f"[ERROR] No se encontró: {archivo}")
    else:
        df = pd.read_csv(archivo)
        generar_graficas_de_barras(df)
        generar_graficas_de_lineas(df)
        generar_graficas_de_pastel(df)
        generar_dispersión(df)
        print("\nTodas las gráficas y libros de Excel fueron generados correctamente en la misma carpeta.")
